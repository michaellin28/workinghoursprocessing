import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import logging
import re

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def read_pos_csv(csv_path):
    """
    Reads specific columns from a POS CSV file until a specific condition is met.

    Args:
        csv_path (str): The path to the POS CSV file.

    Returns:
        pandas.DataFrame: A DataFrame containing 'Name' and 'Work Hours'
                          for relevant rows, or None if an error occurs.
    """
    try:
        # Read the entire CSV first to find the cutoff row
        # Header is in the second row (index 1)
        full_df = pd.read_csv(csv_path, header=1, dtype=str, keep_default_na=False, on_bad_lines='warn')

        cutoff_index = -1
        # Original column B is at index 1
        for index, row in full_df.iterrows():
            if row.iloc[1] == "Role":
                cutoff_index = index
                break

        if cutoff_index == -1:
            logging.warning(f"Cutoff 'Role' not found in column B of {csv_path}. Processing all rows.")
            # If cutoff not found, process all rows read initially
            cutoff_index = len(full_df)

        # Now read the relevant part with correct types, selecting only needed columns
        # Use the determined number of rows (cutoff_index)
        df = pd.read_csv(
            csv_path,
            header=1,
            usecols=["Name", "Work Hours"], # Column A and G
            nrows=cutoff_index,
            # Ensure both columns are read as strings initially for cleaning
            dtype={"Name": str, "Work Hours": str}
        )

        # Clean 'Work Hours' column: remove non-numeric characters and convert
        if "Work Hours" in df.columns:
            # Store original values before cleaning for comparison/logging
            original_hours = df["Work Hours"].copy()
            # Remove any character that is not a digit or a decimal point
            df["Work Hours"] = df["Work Hours"].str.replace(r'[^\d.]', '', regex=True)
            # Convert to numeric, coercing errors (like empty strings after cleaning) to NaN
            df["Work Hours"] = pd.to_numeric(df["Work Hours"], errors='coerce')

            # Identify rows where conversion failed (original was not NaN, but became NaN)
            failed_mask = df["Work Hours"].isna() & original_hours.notna() & (original_hours != '')
            if failed_mask.any():
                 failed_indices = df.index[failed_mask].tolist()
                 logging.warning(f"Could not convert 'Work Hours' for some rows in {csv_path} (indices: {failed_indices}). Original values: {original_hours[failed_mask].tolist()}. These rows will be dropped.")
        else:
            logging.warning(f"'Work Hours' column not found in {csv_path} after initial read.")

        # Drop rows where 'Name' or 'Work Hours' might be NaN after selective reading/conversion
        # Drop rows where 'Name' is missing or 'Work Hours' is NaN (due to read errors or conversion failures)
        df.dropna(subset=["Name", "Work Hours"], inplace=True)

        logging.info(f"Successfully read and processed {len(df)} rows from {csv_path}.")
        return df

    except FileNotFoundError:
        logging.error(f"Error: CSV file not found at {csv_path}")
        return None
    except Exception as e:
        logging.error(f"Error reading or processing CSV file {csv_path}: {e}")
        return None

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def normalize(s: str) -> str:
    """
    Lowercase, replace any non-alphanumeric character with a space,
    then collapse whitespace to single spaces.
    """
    return ' '.join(
        re.sub(r'[^0-9a-z]', ' ', s.strip().lower()).split()
    )

def process_excel(template_path, csv_data, week_choice, output_path):
    if csv_data is None or csv_data.empty:
        return False, "No valid CSV data provided."

    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        if week_choice == "Week 1":
            reg_col, ot_col = 'S', 'T'
        elif week_choice == "Week 2":
            reg_col, ot_col = 'V', 'W'
        else:
            return False, "Invalid week choice. Must be 'Week 1' or 'Week 2'."

        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

        # Build lookup map using normalize()
# Names to ignore (normalized: lowercase, single internal spaces)
        # Use the *normalized* versions of the names to ignore
        ignore_names_set = {normalize("H-R Host"), normalize("Online"), normalize("S-COMMON Server"), normalize("S-Johnny Server")}
        # This will result in: {"h r host", "online", "s common server", "s johnny server"}
        name_to_row = {}
        for row in range(2, ws.max_row + 1):
            val = ws[f'C{row}'].value
            if val:
                name_to_row[normalize(str(val))] = row

        processed = set()

        for _, row in csv_data.iterrows():
            raw_name = str(row['Name'])
            hours = row['Work Hours']

            if not raw_name.strip() or pd.isna(hours):
                logging.warning(f"Skipping invalid row: Name='{raw_name}', Hours={hours}")
                continue

            key = normalize(raw_name)

            # Skip ignored names
            if key in ignore_names_set:
                logging.info(f"Skipping ignored name: '{raw_name}'")
                continue

            if key in processed:
                logging.warning(f"Duplicate entry for '{raw_name}'")
                continue
            processed.add(key)

            reg_hours = min(hours, 40)
            ot_hours  = max(hours - 40, 0)

            target = name_to_row.get(key)
            if target:
                ws[f'{reg_col}{target}'] = reg_hours
                ws[f'{ot_col}{target}'] = ot_hours or ''
                logging.debug(f"Updated {raw_name} (row {target})")
            else:
                # insert below last B‑value row
                last_b = max(r for r in range(2, ws.max_row+1) if ws[f'B{r}'].value)
                ins = last_b + 1
                ws.insert_rows(ins)
                ws[f'B{ins}'].value = raw_name
                ws[f'B{ins}'].fill = red_fill
                ws[f'{reg_col}{ins}'] = reg_hours
                ws[f'{ot_col}{ins}'] = ot_hours or ''
                logging.info(f"Added new '{raw_name}' at row {ins}")

        wb.save(output_path)
        return True, f"Saved to {output_path}"

    except FileNotFoundError:
        logging.error(f"Template not found: {template_path}")
        return False, f"Template not found: {template_path}"
    except Exception as e:
        logging.error(f"Error: {e}")
        return False, f"Error processing: {e}"
    

# Example of how to generate an output filename (optional helper)
def generate_output_filename(template_path, suffix="_processed"):
    """Generates an output filename based on the template name."""
    from pathlib import Path
    p = Path(template_path)
    return p.stem + suffix + p.suffix

# Example Usage (can be commented out or removed)
# if __name__ == "__main__":
#     csv_file = 'path/to/your/pos_data.csv' # Replace with actual path
#     template_file = 'path/to/your/template.xlsx' # Replace with actual path
#     output_dir = 'path/to/your/output/directory' # Replace with actual path
#     week = "Week 1" # Or "Week 2"

#     data = read_pos_csv(csv_file)

#     if data is not None:
#         output_filename = generate_output_filename(template_file, f"_{week.replace(' ', '_')}_output")
#         output_filepath = f"{output_dir}/{output_filename}" # Adjust path joining as needed

#         success, message = process_excel(template_file, data, week, output_filepath)
#         print(message)
#     else:
#         print("Failed to read CSV data.")